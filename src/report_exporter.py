"""
report_exporter.py - CSV and Excel Attendance Report Exporter.

Provides functions to generate:
  1. Standard CSV stream containing attendance details.
  2. Professional Excel workbook (.xlsx) containing 2 sheets:
     - 'Attendance Details': Full record matrix (Date, Time, Period, Student ID, Name, Department, Status).
     - 'Summary': Executive summary metrics (Total Students, Present, Absent, Attendance %, Date Range, Dept, Generated At).
"""

import io
import csv
from datetime import datetime
from typing import Dict, Any, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_attendance_csv(report_data: Dict[str, Any]) -> str:
    """Generate CSV string from report records."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row: Date, Period, Time Slot, Subject, Class Type, Student ID, Student Name, Section, Status, Attendance Time
    writer.writerow(["Date", "Period", "Time Slot", "Subject", "Class Type", "Student ID", "Student Name", "Section", "Status", "Attendance Time"])

    records = report_data.get("all_records") or report_data.get("records") or []
    for r in records:
        writer.writerow([
            r.get("attendance_date", ""),
            f"P{r.get('period_number', '')}",
            r.get("hourly_period", ""),
            r.get("subject", ""),
            r.get("class_type", ""),
            r.get("student_id", ""),
            r.get("student_name", ""),
            r.get("section", ""),
            r.get("status", ""),
            r.get("attendance_time", ""),
        ])

    return output.getvalue()


def generate_attendance_excel(report_data: Dict[str, Any]) -> bytes:
    """Generate Excel binary bytes (.xlsx) with Details and Summary sheets using openpyxl."""
    if not HAS_OPENPYXL:
        return b"Excel export unavailable: openpyxl not installed"
    wb = openpyxl.Workbook()

    # Colors and Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark Slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Soft Green
    present_font = Font(name="Calibri", size=10, color="166534", bold=True)

    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Soft Red
    absent_font = Font(name="Calibri", size=10, color="991B1B", bold=True)

    border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # ── SHEET 1: Attendance Details ──────────────────────────────────────
    ws_details = wb.active
    ws_details.title = "Attendance Details"

    headers = ["Date", "Period", "Time Slot", "Subject", "Class Type", "Student ID", "Student Name", "Section", "Status", "Attendance Time"]
    ws_details.append(headers)

    # Style header row
    for col_num in range(1, len(headers) + 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = cell_border

    records = report_data.get("all_records") or report_data.get("records") or []
    for row_idx, r in enumerate(records, start=2):
        status = (r.get("status") or "ABSENT").upper()

        ws_details.cell(row=row_idx, column=1, value=r.get("attendance_date", "")).alignment = align_center
        ws_details.cell(row=row_idx, column=2, value=f"P{r.get('period_number', '')}").alignment = align_center
        ws_details.cell(row=row_idx, column=3, value=r.get("hourly_period", "")).alignment = align_center
        ws_details.cell(row=row_idx, column=4, value=r.get("subject", "")).alignment = align_left
        ws_details.cell(row=row_idx, column=5, value=r.get("class_type", "")).alignment = align_center
        ws_details.cell(row=row_idx, column=6, value=r.get("student_id", "")).alignment = align_center
        ws_details.cell(row=row_idx, column=7, value=r.get("student_name", "")).alignment = align_left
        ws_details.cell(row=row_idx, column=8, value=r.get("section", "")).alignment = align_center

        status_cell = ws_details.cell(row=row_idx, column=9, value=status)
        status_cell.alignment = align_center

        ws_details.cell(row=row_idx, column=10, value=r.get("attendance_time", "")).alignment = align_center

        if status == "PRESENT":
            status_cell.fill = present_fill
            status_cell.font = present_font
        else:
            status_cell.fill = absent_fill
            status_cell.font = absent_font

        for col_num in range(1, len(headers) + 1):
            ws_details.cell(row=row_idx, column=col_num).border = cell_border

    # Auto-adjust column widths
    for col in ws_details.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_details.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ── SHEET 2: Summary ─────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")

    summary_data = report_data.get("summary") or {}
    ws_summary.append(["HOD Attendance Report Executive Summary"])
    ws_summary.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    ws_summary.append([])

    summary_rows = [
        ("Report Period", f"{summary_data.get('start_date', '')} to {summary_data.get('end_date', '')}"),
        ("Department", summary_data.get("department", "ALL")),
        ("Section", summary_data.get("section", "B")),
        ("Total Registered Students", summary_data.get("total_students", 0)),
        ("Students Present (At least 1 period)", summary_data.get("present_students", 0)),
        ("Students Absent (All periods)", summary_data.get("absent_students", 0)),
        ("Scheduled Periods Evaluated", summary_data.get("total_periods", 0)),
        ("Present Period Slots", summary_data.get("present_periods", 0)),
        ("Absent Period Slots", summary_data.get("absent_periods", 0)),
        ("Overall Attendance Percentage", f"{summary_data.get('attendance_percentage', 0.0)}%"),
        ("Report Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for label, val in summary_rows:
        row_num = ws_summary.max_row + 1
        c1 = ws_summary.cell(row=row_num, column=1, value=label)
        c2 = ws_summary.cell(row=row_num, column=2, value=val)

        c1.font = Font(name="Calibri", size=11, bold=True)
        c1.border = cell_border
        c2.font = Font(name="Calibri", size=11)
        c2.border = cell_border
        c2.alignment = align_left

    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 28

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream.getvalue()


def _clean_pdf_str(val: Any) -> str:
    """Sanitize strings for fpdf2 standard latin-1 core fonts."""
    if val is None:
        return ""
    s = str(val)
    # Replace common typography characters with safe ASCII equivalents
    replacements = {
        "\u2013": "-", "\u2014": "-", "\u2018": "'", "\u2019": "'",
        "\u201c": '"', "\u201d": '"', "\u2022": "*", "\u2026": "...",
        "\xa0": " ",
    }
    for orig, rep in replacements.items():
        s = s.replace(orig, rep)
    return s.encode("latin-1", "replace").decode("latin-1")


def generate_attendance_pdf(report_data: Dict[str, Any]) -> bytes:
    """Generate professional print-ready PDF using fpdf2."""
    from fpdf import FPDF

    summary = report_data.get("summary", {})
    records = report_data.get("all_records") or report_data.get("records") or []

    class PDF(FPDF):
        def header(self):
            self.set_font('helvetica', 'B', 16)
            self.set_text_color(30, 41, 59)
            self.cell(0, 10, 'Classroom AI Attendance Report', border=False, align='C', new_x="LMARGIN", new_y="NEXT")
            self.set_font('helvetica', '', 10)
            self.cell(0, 10, f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', border=False, align='C', new_x="LMARGIN", new_y="NEXT")
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font('helvetica', 'I', 8)
            self.set_text_color(100, 116, 139)
            self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')

    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Summary Section ──────────────────────────────────────────────────
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 8, 'Executive Summary', new_x="LMARGIN", new_y="NEXT")
    pdf.set_font('helvetica', '', 9)

    col_w = 45
    summary_items = [
        ('Department', summary.get('department', 'ALL')),
        ('Section', summary.get('section', 'B')),
        ('Academic Year', summary.get('academic_year', '-')),
        ('Year Level', summary.get('year_level', '-')),
        ('Semester', summary.get('semester', '-')),
        ('Range', f"{summary.get('start_date')} to {summary.get('end_date')}"),
        ('Time Filter', f"{summary.get('from_time', '00:00')} to {summary.get('to_time', '23:59')}"),
        ('Total Students', str(summary.get('total_students', 0))),
        ('Attendance %', f"{summary.get('attendance_percentage', 0.0)}%"),
    ]

    for label, val in summary_items:
        pdf.set_font('helvetica', 'B', 9)
        pdf.cell(30, 6, _clean_pdf_str(f"{label}:"), border=0)
        pdf.set_font('helvetica', '', 9)
        pdf.cell(col_w, 6, _clean_pdf_str(val), border=0)
        if pdf.get_x() > 240: pdf.ln(6)

    pdf.ln(10)

    # ── Table Header ─────────────────────────────────────────────────────
    pdf.set_font('helvetica', 'B', 8)
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)

    headers = [
        ('Date', 20), ('Period', 12), ('Subject', 45), ('Faculty', 35),
        ('Student ID', 25), ('Student Name', 50), ('Sec', 10), ('Status', 20), ('Time', 25)
    ]

    for h, w in headers:
        pdf.cell(w, 8, _clean_pdf_str(h), border=1, align='C', fill=True)
    pdf.ln(8)

    # ── Table Data ───────────────────────────────────────────────────────
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('helvetica', '', 8)

    if not records:
        pdf.cell(242, 10, "No attendance records found for this period/filter.", border=1, align='C')
        pdf.ln(10)
    else:
        for r in records:
            status = (r.get("status") or "ABSENT").upper()

            # Color coding status
            if status == "PRESENT":
                pdf.set_text_color(22, 101, 52) # Dark Green
            else:
                pdf.set_text_color(153, 27, 27) # Dark Red

            pdf.cell(20, 7, _clean_pdf_str(r.get("attendance_date", "")), border=1, align='C')
            pdf.cell(12, 7, _clean_pdf_str(f"P{r.get('period_number', '')}"), border=1, align='C')

            # Trim subject/faculty if too long
            subj = str(r.get("subject", ""))[:25]
            fac = str(r.get("faculty_name", ""))[:20]

            pdf.cell(45, 7, _clean_pdf_str(subj), border=1)
            pdf.cell(35, 7, _clean_pdf_str(fac), border=1)
            pdf.cell(25, 7, _clean_pdf_str(r.get("student_id", "")), border=1, align='C')
            pdf.cell(50, 7, _clean_pdf_str(str(r.get("student_name", ""))[:28]), border=1)
            pdf.cell(10, 7, _clean_pdf_str(r.get("section", "")), border=1, align='C')
            pdf.cell(20, 7, _clean_pdf_str(status), border=1, align='C')
            pdf.cell(25, 7, _clean_pdf_str(r.get("attendance_time", "")), border=1, align='C')
            pdf.ln(7)

    return bytes(pdf.output())
